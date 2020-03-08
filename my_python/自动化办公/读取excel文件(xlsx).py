from openpyxl.reader.excel import load_workbook


"""   # 读一个表里的数据
def readXlsxSheetFile(path):
    #打开excel
    file = load_workbook(filename=path)
    sheets = file.get_sheet_names()  #Excel中所有sheet的名称

    print(sheets)
    #拿出一个sheet的开始读
    sheet = file.get_sheet_by_name(sheets[0])
    print(sheet.max_row)   # 取出该Sheet  共有多少行
    print(sheet.max_column)   # 取出该Sheet  共有多少列
    print(sheet.title)           # 取出该Sheet 的名称

    #读取该sheet的数据
    for rowNum in range(1,sheet.max_row + 1 ):    #循环每一行
        listrow = []
        for columnNum in range(1,sheet.max_column + 1):     #循环每一列
            cell = sheet.cell(row = rowNum,column = columnNum)
            value = cell.value
            listrow.append(value)

        print(listrow)

"""

# 读整个excel里的数据
def readExcelFile(path):
    file = load_workbook(filename=path)
    sheets = file.get_sheet_names()
    excelfile = {}
    for sheetname in sheets:
        sheet = file.get_sheet_by_name(sheetname)

        sheetinfo  = []

        for rowNum in range(1, sheet.max_row + 1) :
            listrow = []
            for columnNum in range(1, sheet.max_column + 1) :
                cell = sheet.cell(row=rowNum, column=columnNum)
                value = cell.value
                listrow.append(value)

            sheetinfo.append(listrow)

        excelfile[sheetname] = sheetinfo

    return excelfile




path = r"C:\Users\Administrator\Desktop\洛阳户籍.xlsx"

# readXlsxSheetFile(path)

print(readExcelFile(path))